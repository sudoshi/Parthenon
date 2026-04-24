#!/usr/bin/env python3
"""
Ingest CMS VSAC value-set expansions into Parthenon's app schema.

Handles two CMS reference files:
  1. dQM VSAC (e.g., dqm_vs_20251117.xlsx) — one sheet, ~224K rows.
  2. Eligible-clinician / hospital eCQM (e.g., ec_hospip_hospop_cms_*.xlsx)
     — one sheet per CMS measure (CMS2v15, ...).

Populates:
  - app.vsac_value_sets              (deduped by value_set_oid)
  - app.vsac_value_set_codes         (UPSERT on oid+code+system)
  - app.vsac_measures                (one row per CMS measure sheet)
  - app.vsac_measure_value_sets      (M2M measure → value set)

Run with ~/.pgpass for `claude_dev` on host PG17.

Usage:
  python3 scripts/importers/ingest_vsac.py dqm_vs_20251117.xlsx
  python3 scripts/importers/ingest_vsac.py ec_hospip_hospop_cms_20250508.xlsx
  python3 scripts/importers/ingest_vsac.py --both   # both files from repo root
"""

from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from psycopg2.extras import Json, execute_values

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DQM = REPO_ROOT / "dqm_vs_20251117.xlsx"
DEFAULT_EC = REPO_ROOT / "ec_hospip_hospop_cms_20250508.xlsx"

DSN_DEFAULT = "host=127.0.0.1 dbname=parthenon user=claude_dev"

# Column name → normalized key. Column order varies between the two files.
FIELD_MAP = {
    "CMS ID": "cms_id",
    "Program Candidate": "program_candidate",
    "CBE Number": "cbe_number",
    "Value Set Name": "name",
    "Value Set OID": "value_set_oid",
    "QDM Category": "qdm_category",
    "Definition Version": "definition_version",
    "Expansion Version": "expansion_version",
    "Purpose: Clinical Focus": "purpose_clinical_focus",
    "Purpose: Data Element Scope": "purpose_data_scope",
    "Purpose: Inclusion Criteria": "purpose_inclusion",
    "Purpose: Exclusion Criteria": "purpose_exclusion",
    "Code": "code",
    "Description": "description",
    "Code System": "code_system",
    "Code System OID": "code_system_oid",
    "Code System Version": "code_system_version",
    "Expansion ID": "expansion_id",
}

log = logging.getLogger("ingest_vsac")


@dataclass(frozen=True)
class SheetScan:
    value_set_rows: dict[str, dict]       # oid → value-set metadata
    code_rows: list[tuple]                # (oid, code, description, sys, sys_oid, sys_ver)
    measure_row: dict | None              # per-sheet measure metadata
    measure_value_sets: set[str]          # oids belonging to this measure


def _find_header_row(ws) -> tuple[int, list[str]]:
    """Locate the row whose first cell is 'CMS ID' or 'Value Set Name'."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > 10:
            break
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not cells:
            continue
        first = cells[0]
        if first in ("CMS ID", "Value Set Name"):
            return row_idx, cells
    raise ValueError(f"Could not find header row in sheet {ws.title!r}")


def scan_sheet(ws, source_file: str) -> SheetScan:
    header_row, headers = _find_header_row(ws)
    idx_by_key = {FIELD_MAP[h]: i for i, h in enumerate(headers) if h in FIELD_MAP}

    def get(row: tuple, key: str) -> str | None:
        i = idx_by_key.get(key)
        if i is None or i >= len(row):
            return None
        v = row[i]
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    value_sets: dict[str, dict] = {}
    codes: list[tuple] = []
    measure_row: dict | None = None
    measure_oids: set[str] = set()
    seen_codes: set[tuple[str, str, str]] = set()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        oid = get(row, "value_set_oid")
        if not oid:
            continue
        name = get(row, "name")
        if name is None:
            continue

        if oid not in value_sets:
            value_sets[oid] = {
                "value_set_oid": oid,
                "name": name[:500],
                "definition_version": get(row, "definition_version"),
                "expansion_version": get(row, "expansion_version"),
                "expansion_id": get(row, "expansion_id"),
                "qdm_category": get(row, "qdm_category"),
                "purpose_clinical_focus": get(row, "purpose_clinical_focus"),
                "purpose_data_scope": get(row, "purpose_data_scope"),
                "purpose_inclusion": get(row, "purpose_inclusion"),
                "purpose_exclusion": get(row, "purpose_exclusion"),
                "source_file": source_file,
            }

        code = get(row, "code")
        code_sys = get(row, "code_system")
        if code and code_sys:
            key = (oid, code, code_sys)
            if key not in seen_codes:
                seen_codes.add(key)
                codes.append((
                    oid, code[:100],
                    (get(row, "description") or "")[:4000] or None,
                    code_sys[:80],
                    get(row, "code_system_oid"),
                    get(row, "code_system_version"),
                ))

        cms_id = get(row, "cms_id")
        if cms_id:
            if measure_row is None:
                measure_row = {
                    "cms_id": cms_id,
                    "cbe_number": get(row, "cbe_number"),
                    "program_candidate": get(row, "program_candidate"),
                    "title": None,
                    "expansion_version": get(row, "expansion_version"),
                }
            measure_oids.add(oid)

    return SheetScan(value_sets, codes, measure_row, measure_oids)


def iter_sheets(path: Path) -> Iterator[tuple[str, object]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    for name in wb.sheetnames:
        yield name, wb[name]


def upsert_value_sets(cur, rows: list[dict]) -> None:
    if not rows:
        return
    execute_values(
        cur,
        """
        INSERT INTO app.vsac_value_sets (
            value_set_oid, name, definition_version, expansion_version,
            expansion_id, qdm_category, purpose_clinical_focus,
            purpose_data_scope, purpose_inclusion, purpose_exclusion,
            source_files, ingested_at
        ) VALUES %s
        ON CONFLICT (value_set_oid) DO UPDATE SET
            name = EXCLUDED.name,
            definition_version = COALESCE(EXCLUDED.definition_version, app.vsac_value_sets.definition_version),
            expansion_version = COALESCE(EXCLUDED.expansion_version, app.vsac_value_sets.expansion_version),
            expansion_id = COALESCE(EXCLUDED.expansion_id, app.vsac_value_sets.expansion_id),
            qdm_category = COALESCE(EXCLUDED.qdm_category, app.vsac_value_sets.qdm_category),
            purpose_clinical_focus = COALESCE(EXCLUDED.purpose_clinical_focus, app.vsac_value_sets.purpose_clinical_focus),
            purpose_data_scope = COALESCE(EXCLUDED.purpose_data_scope, app.vsac_value_sets.purpose_data_scope),
            purpose_inclusion = COALESCE(EXCLUDED.purpose_inclusion, app.vsac_value_sets.purpose_inclusion),
            purpose_exclusion = COALESCE(EXCLUDED.purpose_exclusion, app.vsac_value_sets.purpose_exclusion),
            source_files = (
                SELECT to_jsonb(array_agg(DISTINCT f))
                FROM jsonb_array_elements_text(
                    app.vsac_value_sets.source_files || EXCLUDED.source_files
                ) AS f
            ),
            ingested_at = NOW()
        """,
        [
            (
                r["value_set_oid"], r["name"], r["definition_version"],
                r["expansion_version"], r["expansion_id"], r["qdm_category"],
                r["purpose_clinical_focus"], r["purpose_data_scope"],
                r["purpose_inclusion"], r["purpose_exclusion"],
                Json([r["source_file"]]),
            )
            for r in rows
        ],
        template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())",
        page_size=500,
    )


def upsert_codes(cur, rows: list[tuple]) -> None:
    if not rows:
        return
    execute_values(
        cur,
        """
        INSERT INTO app.vsac_value_set_codes
            (value_set_oid, code, description, code_system, code_system_oid, code_system_version)
        VALUES %s
        ON CONFLICT (value_set_oid, code, code_system) DO UPDATE SET
            description = COALESCE(EXCLUDED.description, app.vsac_value_set_codes.description),
            code_system_oid = COALESCE(EXCLUDED.code_system_oid, app.vsac_value_set_codes.code_system_oid),
            code_system_version = COALESCE(EXCLUDED.code_system_version, app.vsac_value_set_codes.code_system_version)
        """,
        rows,
        page_size=2000,
    )


def upsert_measure(cur, measure: dict) -> None:
    cur.execute(
        """
        INSERT INTO app.vsac_measures
            (cms_id, cbe_number, program_candidate, title, expansion_version, ingested_at)
        VALUES (%s, %s, %s, %s, %s, NOW())
        ON CONFLICT (cms_id) DO UPDATE SET
            cbe_number = COALESCE(EXCLUDED.cbe_number, app.vsac_measures.cbe_number),
            program_candidate = COALESCE(EXCLUDED.program_candidate, app.vsac_measures.program_candidate),
            expansion_version = COALESCE(EXCLUDED.expansion_version, app.vsac_measures.expansion_version),
            ingested_at = NOW()
        """,
        (
            measure["cms_id"], measure["cbe_number"], measure["program_candidate"],
            measure["title"], measure["expansion_version"],
        ),
    )


def upsert_measure_links(cur, cms_id: str, oids: set[str]) -> None:
    if not oids:
        return
    execute_values(
        cur,
        """
        INSERT INTO app.vsac_measure_value_sets (cms_id, value_set_oid)
        VALUES %s
        ON CONFLICT (cms_id, value_set_oid) DO NOTHING
        """,
        [(cms_id, oid) for oid in oids],
        page_size=500,
    )


def ingest_file(path: Path, conn) -> dict:
    if not path.is_file():
        raise FileNotFoundError(path)

    source_file = path.name
    log.info("ingesting %s", source_file)
    totals = {"value_sets": 0, "codes": 0, "measures": 0, "links": 0, "sheets": 0}

    with conn.cursor() as cur:
        for sheet_name, ws in iter_sheets(path):
            try:
                scan = scan_sheet(ws, source_file)
            except ValueError as e:
                log.warning("skip sheet %s: %s", sheet_name, e)
                continue

            totals["sheets"] += 1
            if scan.value_set_rows:
                upsert_value_sets(cur, list(scan.value_set_rows.values()))
                totals["value_sets"] += len(scan.value_set_rows)
            if scan.code_rows:
                upsert_codes(cur, scan.code_rows)
                totals["codes"] += len(scan.code_rows)
            if scan.measure_row:
                upsert_measure(cur, scan.measure_row)
                upsert_measure_links(cur, scan.measure_row["cms_id"], scan.measure_value_sets)
                totals["measures"] += 1
                totals["links"] += len(scan.measure_value_sets)

            if totals["sheets"] % 10 == 0:
                conn.commit()
                log.info("  checkpoint: %s", totals)

    conn.commit()
    log.info("done: %s", totals)
    return totals


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("file", nargs="?", help="xlsx file to ingest")
    p.add_argument("--both", action="store_true", help="ingest both default files from repo root")
    p.add_argument("--dsn", default=DSN_DEFAULT)
    args = p.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    paths: list[Path] = []
    if args.both:
        paths = [DEFAULT_DQM, DEFAULT_EC]
    elif args.file:
        paths = [Path(args.file)]
    else:
        p.print_help()
        return 2

    conn = psycopg2.connect(args.dsn)
    try:
        for path in paths:
            ingest_file(path, conn)
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
